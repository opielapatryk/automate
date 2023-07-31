# this app automate file editing in my job
import openpyxl

# otwórz plik Excel
wb = openpyxl.load_workbook('stany.xlsx')

# wybierz arkusz
sheet = wb.active

# zmień wartości w kolumnie C
for cell in sheet['C']:
    if cell.value == 'GS':
        cell.value = 'Schneider Electric'

# usuń kolumny G do Z
sheet.delete_cols(7, 20)

# przenieś zawartość kolumny F do kolumny O
for row in sheet.iter_rows(min_row=1, min_col=6, max_col=6):
    for cell in row:
        sheet.cell(row=cell.row, column=17).value = cell.value

# usun kolumne F
sheet.delete_cols(6)

# przenieś zawartość kolumny A do kolumny Q
for row in sheet.iter_rows(min_row=1, max_col=1):
    for cell in row:
        sheet.cell(row=cell.row, column=18).value = cell.value

# usun kolumne A
sheet.delete_cols(1)

# wypełnij kolumnę H
sheet['H1'] = 'Stan towaru'
for cell in sheet['H2:H' + str(sheet.max_row)]:
    cell[0].value = 'New Sealed'

# zapisz plik Excel
wb.save('stany.xlsx')